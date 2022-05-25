import os
import tkinter as tk
from datetime import datetime, timedelta
from tkinter import *
from tkinter import filedialog
import openpyxl as ex
import pandas as pd
from tkcalendar import DateEntry

# Widget Layout
root = tk.Tk()

root.geometry("400x210")
root.title('Santos Productivity')
Label(root, text="Click the Button to browse the files").place(x=26, y=30)  # Label
Label(root, text="* .csv file only", font=('Arial', 10)).place(x=285, y=60)  # Label

# Start date range
Label(root, text="Start date").place(x=26, y=90)
startdate = DateEntry(root, locale='en_US', date_pattern='MM/dd/yyyy')
startdate.place(x=26, y=120)

# End date range
Label(root, text="End date").place(x=235, y=90)
enddate = DateEntry(root, locale='en_US', date_pattern='MM/dd/yyyy')
enddate.place(x=235, y=120)

root.file_name = ''


# opening the file
def open_file():
    root.file_name = filedialog.askopenfilename(parent=root, initialdir="/",
                                                filetypes=(("CSV Files", "*.csv"), ("All", "*.*")))


# choose file button
Button(root, text="Choose File", command=open_file).place(x=270, y=26)


# Time stamp convertor
def timestamp_cov(date):
    try:
        datetime.strptime(date, "%m/%d/%Y %I:%M %p")
        data = "%m/%d/%Y %I:%M %p"
    except:
        data = "%m/%d/%y %H:%M"

    return datetime.timestamp(datetime.strptime(date, data))


# Timestamp to Human read
def ts_2_hrd(date):
    return datetime.fromtimestamp(date).strftime('%Y-%m-%d')


def main():
    # reading the file
    df = pd.read_csv(root.file_name)

    # Remove the spaces in CSV
    df = df[df['Tested On'].notna()]

    # Improper date to timestamp
    df['Timestamp'] = df.apply(lambda row: timestamp_cov(row['Tested On']), axis=1)

    # Timestamp to standard Human reading date format
    df['Human read datetime'] = df.apply(lambda row: ts_2_hrd(row['Timestamp']), axis=1)

    # Date getting from the input

    # start date same date as per give input
    start_date = str(startdate.get_date())

    # For here converting to the next date ( Eg: Given date = next date) this will give the proper output
    e_date = (enddate.get_date())
    td = timedelta(1)
    end_date = str(e_date + td)

    # sorting the data's as per given input
    filtered_df = df.loc[(df['Human read datetime'] >= start_date) & (df['Human read datetime'] < end_date)]

    # made the Excel format of the sorted file
    sort = pd.DataFrame(filtered_df)

    # Removing the improper date & timestamp column
    sort.pop('Timestamp')
    sort.pop('Tested On')
    # Exporting the data to Excel format to perform the next steps
    sort.to_excel("temp_file.xlsx")

    # export as the groupby
    '''print(sort.groupby(['Tested By']).count())'''

    # Sum of the total count of the case execution
    '''print((sort.groupby(['Tested By']).count()).sum())'''

    # Loading workbook to the WB as local storage
    wb = ex.load_workbook("temp_file.xlsx")

    # Removing the temporary
    os.remove("temp_file.xlsx")

    sheet = wb.active
    sheet.cell(row=1, column=1)

    max_row = sheet.max_row

    res, file_ = {}, []

    # Loop will print all rows name
    for i in range(2, max_row + 1):
        cell_obj_1 = sheet.cell(row=i, column=2)
        cell_obj_2 = sheet.cell(row=i, column=3)

        file_.append((cell_obj_1.value, cell_obj_2.value))

    # print(f" length of file: {len(file_)}")

    def group_excel_by_date(file: list):

        list_rows = file

        for item in list_rows:
            # if the name hasn't been entered to dict yet
            if item[0] not in res:
                res[item[0]] = {item[1]: 1}
            elif item[0] in res:
                if item[1] in res[item[0]]:
                    res[item[0]][item[1]] = res[item[0]][item[1]] + 1
                else:
                    res[item[0]][item[1]] = 1

        return res

    result = group_excel_by_date(file_)

    df = pd.DataFrame(result).T
    df = df.reindex(sorted(df.columns), axis=1)
    file = filedialog.asksaveasfilename(defaultextension=".xlsx")
    df.to_excel(str(file))
    # print("Program completed")


Button(root, text="Save", command=main).place(x=302, y=170)

root.mainloop()
